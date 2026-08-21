import os
import asyncio
import requests
import nest_asyncio
import re
import urllib3
import io
import smtplib
import pandas as pd
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from email.message import EmailMessage

# SSL 경고 메세지 무시
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
nest_asyncio.apply()

def get_dow_jones_sso_url():
    """requests를 사용하여 로그인 버튼 클릭을 시뮬레이션하고 SSO 리다이렉트 URL을 가져옵니다."""
    session = requests.Session()
    session.verify = False 
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })
    
    target_url = "https://cma.opisnet.com" 
    try:
        print(f"[{target_url}] 접속하여 SSO 리다이렉트 URL 확인 중...")
        response = session.get(target_url, timeout=30)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        form = soup.find('form')
        if not form:
            return None

        payload = {}
        for hidden_input in form.find_all('input', type='hidden'):
            name = hidden_input.get('name')
            value = hidden_input.get('value', '')
            if name:
                payload[name] = value
                
        payload['dow_jones_idp'] = 'Log in with Dow Jones'
        
        action_url = form.get('action')
        if not action_url or action_url == '/':
            action_url = target_url
        elif action_url.startswith('/'):
            action_url = target_url.rstrip('/') + action_url

        post_response = session.post(action_url, data=payload, allow_redirects=False, timeout=30)
        
        if post_response.status_code in (301, 302, 303, 307):
            return post_response.headers.get('Location')
    except Exception as e:
        print(f"⚠️ SSO URL 획득 실패: {e}")
    return None

async def get_links_and_cookies_with_retry(max_retries=5):
    attempt = 0
    while attempt < max_retries:
        attempt += 1
        print(f"\n[{attempt}/{max_retries}] 전체 프로세스 시도 중 (브라우저 재시작 포함)...")
        
        try:
            # 1. requests로 SSO URL 획득
            sso_url = get_dow_jones_sso_url()
            if not sso_url:
                print("   ⚠️ SSO URL을 가져오지 못했습니다. 잠시 후 다시 시도합니다.")
                await asyncio.sleep(5)
                if attempt < max_retries: continue
                else: raise Exception("SSO URL 획득 최종 실패")

            # 2. 브라우저 실행 단계
            async with async_playwright() as p:
                print(f"브라우저를 실행합니다 (headless=True)...")
                browser = await p.chromium.launch(headless=True)
                context = await browser.new_context()
                page = await context.new_page()

                # 3. 획득한 SSO URL로 직접 접속
                print(f"SSO 페이지로 직접 접속합니다...")
                await page.goto(sso_url, wait_until="domcontentloaded", timeout=60000)
                
                # 4. 로그인 정보 입력
                cma_user = os.environ.get("CMA_USER")
                cma_password = os.environ.get("CMA_PASSWORD")

                email_selector = 'input[name="emailOrUserID"], #email'
                await page.wait_for_selector(email_selector, state="visible", timeout=30000)
                
                print("계정 정보 입력 중...")
                await page.fill(email_selector, cma_user)
                password_selector = 'input[type="password"], #password-form-item'
                await page.wait_for_selector(password_selector, state="visible", timeout=15000)
                await page.fill(password_selector, cma_password)
                await page.press(password_selector, 'Enter')

                print("로그인 완료 대기 중...")
                await page.wait_for_url(lambda url: "cma.opisnet.com" in url and "login" not in url.lower(), timeout=45000)
                
                # 5. 목록 페이지 이동 (Daily / Monthly 추출)
                list_url = "https://cma.opisnet.com/publications/market-advisory-service?page=1&itemsPerPage=100"
                print(f"Daily/Monthly 목록 페이지 로딩 중... ({list_url})")
                await page.goto(list_url, wait_until="domcontentloaded", timeout=60000)
                
                await page.wait_for_selector('a:has-text("Daily North America")', timeout=30000)
                await page.wait_for_timeout(2000)

                print("Daily 및 Monthly 보고서 링크 추출 중...")
                daily_node = page.locator('a:has-text("Daily North America")').first
                monthly_node = page.locator('a:has-text("North America Aromatics - Benzene Contract Price")').first
                
                daily_url = await daily_node.get_attribute("href")
                monthly_url = await monthly_node.get_attribute("href")

                # 6. Weekly 전용 검색 페이지 이동 및 선별
                weekly_search_url = "https://cma.opisnet.com/find?search_api_fulltext=Global%20Aromatics%20-%20Weekly%20Market%20Report"
                print(f"Weekly 전용 검색 페이지 로딩 중... ({weekly_search_url})")
                await page.goto(weekly_search_url, wait_until="domcontentloaded", timeout=60000)
                
                await page.wait_for_selector('a:has-text("Global Aromatics - Weekly Market Report")', timeout=30000)
                await page.wait_for_timeout(2000)

                print("Weekly 보고서 후보군 탐색 및 최신 링크 선별 중...")
                weekly_locators = await page.locator('a:has-text("Global Aromatics - Weekly Market Report")').all()
                weekly_candidates = []
                for loc in weekly_locators[:10]:
                    text = await loc.text_content()
                    href = await loc.get_attribute("href")
                    if href and text:
                        issue_match = re.search(r'Issue\s*(\d+)', text, re.IGNORECASE)
                        issue_num = int(issue_match.group(1)) if issue_match else 0
                        is_reissue = 1 if ('reissue' in text.lower() or 'revised' in text.lower()) else 0
                        
                        weekly_candidates.append({
                            'text': text.strip(),
                            'href': href,
                            'issue_num': issue_num,
                            'is_reissue': is_reissue
                        })
                
                if weekly_candidates:
                    # 1순위: issue_num 내림차순, 2순위: is_reissue(1) 우선
                    weekly_candidates.sort(key=lambda x: (x['issue_num'], x['is_reissue']), reverse=True)
                    best_weekly = weekly_candidates[0]
                    weekly_url = best_weekly['href']
                    print(f"📌 선택된 Weekly 리포트: '{best_weekly['text']}' (Issue: {best_weekly['issue_num']}, Reissue: {bool(best_weekly['is_reissue'])})")
                else:
                    weekly_node = page.locator('a:has-text("Global Aromatics - Weekly Market Report")').first
                    weekly_url = await weekly_node.get_attribute("href")
                
                base_url = "https://cma.opisnet.com"
                data = {
                    "daily_url": base_url + daily_url if daily_url.startswith('/') else daily_url,
                    "weekly_url": base_url + weekly_url if weekly_url.startswith('/') else weekly_url,
                    "monthly_url": base_url + monthly_url if monthly_url.startswith('/') else monthly_url,
                    "cookies": await context.cookies()
                }
                
                await browser.close()
                return data

        except Exception as e:
            print(f"⚠️ {attempt}회차 시도 중 에러 발생: {e}")
            if attempt < max_retries:
                wait_time = 5 * attempt
                print(f"{wait_time}초 후 다시 시도합니다...")
                await asyncio.sleep(wait_time)
            else:
                print("❌ 모든 재시도 횟수를 초과했습니다.")
                raise e

def convert_to_yyyymmdd(text):
    """다양한 날짜 형식을 'YYYYMMDD'로 변환"""
    date_match = re.search(r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', str(text))
    if date_match:
        try: return datetime.strptime(date_match.group(1), '%d %b %Y').strftime('%Y%m%d')
        except: pass
    iso_match = re.search(r'(\d{4}-\d{2}-\d{2})', str(text))
    if iso_match:
        try: return datetime.strptime(iso_match.group(1), '%Y-%m-%d').strftime('%Y%m%d')
        except: pass
    return str(text)[:10]

def apply_excel_style(ws):
    """시트 스타일 적용"""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            else:
                if ws.title == 'Summary':
                    if cell.column == 3:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '0.0'
                        except: pass
                    # 기준 날짜(Col 4) 및 실제 날짜(Col 5) 숫자 서식 적용
                    if cell.column in (4, 5):
                        try:
                            cell.value = int(str(cell.value).strip())
                            cell.number_format = '0'
                        except: pass

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

def check_monthly_business_day(now):
    """
    당월 첫 1~5영업일 또는 마지막 1~3영업일인지 확인하고,
    해당할 경우 (True, '기준날짜(YYYYMM01)')를 반환합니다.
    """
    start_of_month = now.replace(day=1)
    if now.month == 12:
        next_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        next_month = now.replace(month=now.month + 1, day=1)
    end_of_month = next_month - timedelta(days=1)

    # 당월 모든 평일(월~금) 목록 생성
    b_days = pd.bdate_range(start=start_of_month, end=end_of_month).date

    first_5_bdays = b_days[:5]   # 당월 첫 1~5영업일
    last_3_bdays = b_days[-3:]   # 당월 마지막 1~3영업일

    today_date = now.date()

    if today_date in last_3_bdays:
        # 월말 마지막 3영업일 -> 익월 1일
        ref_date = f"{next_month.year}{next_month.month:02d}01"
        return True, ref_date
    elif today_date in first_5_bdays:
        # 월초 첫 5영업일 -> 당월 1일
        ref_date = f"{now.year}{now.month:02d}01"
        return True, ref_date
    else:
        return False, None

def process_data(data):
    session = requests.Session()
    session.verify = False 
    for cookie in data['cookies']:
        session.cookies.set(name=cookie['name'], value=cookie['value'], domain=cookie['domain'], path=cookie['path'])
    
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })

    now = datetime.now()
    current_weekday = now.weekday()  # 0: 월, 1: 화, 2: 수, 3: 목, 4: 금, 5: 토, 6: 일

    # 실행 조건 판별
    run_weekly = current_weekday in (0, 4)  # 월요일(0) 또는 금요일(4)
    run_monthly, monthly_ref_date = check_monthly_business_day(now)  # 월말 3영업일 또는 월초 5영업일

    summary_data = {
        "daily_means": [], 
        "daily_date": None, 
        "daily_actual_date": None,
        "weekly_mean": None, 
        "weekly_date": None, 
        "weekly_actual_date": None,
        "monthly_cents": None, 
        "monthly_date": None,
        "monthly_actual_date": None
    }

    # 1. Daily (항상 실행)
    print("\n" + "="*20 + " [1] Daily 분석 시작 " + "="*20)
    d_resp = session.get(data['daily_url'])
    d_soup = BeautifulSoup(d_resp.text, 'html.parser')
    page_text_daily = d_soup.get_text(separator='\n', strip=True)
    try:
        parsed_daily_date = convert_to_yyyymmdd(page_text_daily)
        summary_data["daily_date"] = parsed_daily_date
        summary_data["daily_actual_date"] = parsed_daily_date
        
        benzene_start = re.search(r'Benzene\s*\(Houston,\s*TX\s*basis\)', page_text_daily, re.IGNORECASE)
        if benzene_start:
            section = page_text_daily[benzene_start.start():]
            pattern = r'(?<=Cents per gallon)(.*?)(?=USD per metric ton \(converted\))'
            match = re.search(pattern, section, re.DOTALL | re.IGNORECASE)
            if match:
                raw_text = match.group(1).strip()
                rows = []
                for line in raw_text.split('\n'):
                    m = re.search(r'^([A-Za-z]+)\s+([\d\.,]+)\s+to\s+([\d\.,]+)', line.strip())
                    if m: rows.append([m.group(1), float(m.group(2).replace(',', '')), float(m.group(3).replace(',', ''))])
                if rows:
                    benzene_df = pd.DataFrame(rows, columns=['Month', 'Low', 'High'])
                    benzene_df['Mean'] = (benzene_df['Low'] + benzene_df['High']) / 2
                    summary_data["daily_means"] = benzene_df['Mean'].tolist()
                    print(f"📅 Daily Date (기준/실제): {summary_data['daily_date']}")
                    print(benzene_df)
    except Exception as e: 
        print(f"⚠️ Daily 에러: {e}")

    # 2. Weekly (금/월 실행)
    if run_weekly:
        print("\n" + "="*20 + " [2] Weekly 분석 시작 (금/월 실행) " + "="*20)
        try:
            # 기준 날짜 계산: 금요일이면 당일, 월요일이면 -3일(금요일)
            if current_weekday == 4:
                summary_data["weekly_date"] = now.strftime('%Y%m%d')
            elif current_weekday == 0:
                summary_data["weekly_date"] = (now - timedelta(days=3)).strftime('%Y%m%d')

            w_resp = session.get(data['weekly_url'])
            w_soup = BeautifulSoup(w_resp.text, 'html.parser')
            excel_url = None
            for a in w_soup.find_all('a', href=True):
                if ".xlsx" in a['href'] or ".xlsx" in a.get_text():
                    excel_url = "https://cma.opisnet.com" + a['href'] if a['href'].startswith('/') else a['href']
                    break
            if excel_url:
                e_resp = session.get(excel_url)
                if e_resp.status_code == 200:
                    df_raw = pd.read_excel(io.BytesIO(e_resp.content), header=None)
                    row_market, row_type = -1, -1
                    for r in range(min(20, df_raw.shape[0])):
                        cell_val = str(df_raw.iloc[r, 0]).strip().upper()
                        if "MARKET" in cell_val: row_market = r
                        if "TYPE" in cell_val: row_type = r
                    if row_market != -1:
                        target_col = -1
                        for col in range(df_raw.shape[1]):
                            if "Benzene" in str(df_raw.iloc[row_market, col]) and "Spot" in str(df_raw.iloc[row_type, col]):
                                target_col = col; break
                        if target_col != -1:
                            final_df = pd.concat([df_raw.iloc[0:9, [0, target_col, target_col+1]], df_raw.iloc[-2:, [0, target_col, target_col+1]]])
                            final_df = final_df.reset_index(drop=True)
                            final_df.columns = range(final_df.shape[1])
                            def calculate_mean(row):
                                try: return (float(row[1]) + float(row[2])) / 2
                                except: return None
                            final_df[3] = final_df.apply(calculate_mean, axis=1)
                            summary_data["weekly_mean"] = final_df.iloc[-1, 3]
                            
                            # 문서 내 실제 기재 날짜 추출
                            summary_data["weekly_actual_date"] = convert_to_yyyymmdd(final_df.iloc[-1, 0])
                            
                            print(f"📅 Weekly 기준 날짜(고정): {summary_data['weekly_date']}, 실제 날짜: {summary_data['weekly_actual_date']}")
                            print(final_df)
        except Exception as e: 
            print(f"⚠️ Weekly 에러: {e}")
    else:
        print("\n⏩ [2] Weekly 분석 건너뜀 (금/월 아님)")

    # 3. Monthly (월말 3영업일 또는 월초 5영업일 실행)
    if run_monthly:
        print("\n" + "="*20 + " [3] Monthly 분석 시작 (월말 3영업일 / 월초 5영업일 실행) " + "="*20)
        try:
            summary_data["monthly_date"] = monthly_ref_date

            m_resp = session.get(data['monthly_url'])
            m_soup = BeautifulSoup(m_resp.text, 'html.parser')
            page_text_monthly = m_soup.get_text(separator='\n', strip=True)
            
            # 본문 내 실제 날짜 추출
            summary_data["monthly_actual_date"] = convert_to_yyyymmdd(page_text_monthly)

            # 패턴 매칭
            price_match = re.search(r'settlement price of\s*\$(\d+(?:\.\d+)?)\s*per gallon', page_text_monthly, re.IGNORECASE)
            if not price_match:
                price_match = re.search(r'average value of\s*\$(\d+(?:\.\d+)?)\s*per gallon', page_text_monthly, re.IGNORECASE)
                
            if price_match:
                summary_data["monthly_cents"] = round(float(price_match.group(1)) * 100, 2)
                print(f"📅 Monthly 기준 날짜(고정): {summary_data['monthly_date']}, 실제 날짜: {summary_data['monthly_actual_date']}")
                print(f"💰 CP: {summary_data['monthly_cents']} cents")
            else:
                print("⚠️ Monthly 가격 패턴을 찾을 수 없습니다. (리포트 문구 변경 의심)")
        except Exception as e: 
            print(f"⚠️ Monthly 에러: {e}")
    else:
        print("\n⏩ [3] Monthly 분석 건너뜀 (월말 3영업일 / 월초 5영업일 기간 아님)")

    # 4. Final Excel 생성 (동적 Row 구성 + 실제 날짜 열)
    print("\n" + "="*20 + " [4] Excel 보고서 생성 " + "="*20)
    today_str = now.strftime('%Y-%m-%d')
    quot_no_list = ["60M60681", "60M60682", "60M60683", "60M60686", "60M60684"]
    marker_names = ["US BZ DDP Spot Daily(M월)", "US BZ DDP Spot Daily(M+1월)", "US BZ DDP Spot Daily(M+2월)", "US BZ DDP Spot Weekly", "US BZ Monthly Contract Price(CP) cent/gal"]
    
    final_rows = []
    
    # Daily 3개 행 (항상 추가)
    for i in range(3):
        val = summary_data["daily_means"][i] if i < len(summary_data["daily_means"]) else "N/A"
        final_rows.append([marker_names[i], quot_no_list[i], val, summary_data["daily_date"], summary_data["daily_actual_date"]])
    
    # Weekly 행 (금/월에만 추가)
    if run_weekly:
        val = summary_data["weekly_mean"] if summary_data["weekly_mean"] is not None else "N/A"
        final_rows.append([marker_names[3], quot_no_list[3], val, summary_data["weekly_date"], summary_data["weekly_actual_date"]])
        
    # Monthly 행 (월말 3영업일/월초 5영업일에만 추가)
    if run_monthly:
        val = summary_data["monthly_cents"] if summary_data["monthly_cents"] is not None else "N/A"
        final_rows.append([marker_names[4], quot_no_list[4], val, summary_data["monthly_date"], summary_data["monthly_actual_date"]])

    # DataFrame 생성
    final_summary_df = pd.DataFrame(final_rows, columns=['Marker 가격', 'Quot. No', today_str, '기준 날짜', '실제 날짜'])
    
    # URL 시트 목록 동적 생성
    url_rows = [["Daily", data["daily_url"]]]
    if run_weekly:
        url_rows.append(["Weekly", data["weekly_url"]])
    if run_monthly:
        url_rows.append(["Monthly", data["monthly_url"]])
    url_df = pd.DataFrame(url_rows, columns=["Category", "URL"])

    xlsx_file_name = f"CMA_OPIS_{now.strftime('%Y%m%d')}.xlsx"
    with pd.ExcelWriter(xlsx_file_name, engine='openpyxl') as writer:
        final_summary_df.to_excel(writer, sheet_name='Summary', index=False)
        url_df.to_excel(writer, sheet_name='URL', index=False)
        workbook = writer.book
        for sheet_name in ['Summary', 'URL']:
            ws = workbook[sheet_name]
            apply_excel_style(ws)
            if sheet_name == 'URL':
                for i, row_item in enumerate(url_rows, start=2):
                    cell = ws.cell(row=i, column=2)
                    cell.hyperlink = row_item[1]
                    cell.font = Font(color="0000FF", underline="single")

    print(f"💾 저장 완료: {xlsx_file_name}")
    print("\n" + "★"*20 + " [최종 요약 결과] " + "★"*20)
    print(final_summary_df)
    print("★"*57 + "\n")
    
    return xlsx_file_name, final_summary_df

async def main():
    try:
        data = await get_links_and_cookies_with_retry(max_retries=5)
        file_name, df_cma_result = process_data(data)
        
        today_str = datetime.now().strftime('%Y-%m-%d')

        print("=== 메일 발송 준비 ===")

        sender_email = os.environ.get("GMAIL_USER")
        app_password = os.environ.get("GMAIL_APP_PASSWORD")

        to_emails = ["rchangjo@sk.com", "hyo548@sk.com"]
        cc_emails = ["jp_lee@sk.com"]

        subject = f"CMA {today_str}"

        html_table = df_cma_result.to_html(justify='center', index=False)
        custom_table_tag = '<table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse; text-align:center; font-family:Calibri, Arial, sans-serif; font-size:13px;">'
        html_table = html_table.replace('<table border="1" class="dataframe">', custom_table_tag)

        html_body = f"""
        <html>
        <body style="margin:0; padding:0;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
                <tr>
                    <td style="padding:20px; font-family:Calibri, Arial, sans-serif; font-size:14px; color:#000000;">
                        안녕하세요,<br><br>
                        오늘자 CMA 추출 결과입니다.<br>
                        상세 내용은 첨부파일을 확인해 주시기 바랍니다.<br><br>
                        {html_table}
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """

        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = ", ".join(to_emails)
        msg['Cc'] = ", ".join(cc_emails)
        msg.set_content("HTML 뷰어를 지원하는 메일 클라이언트를 사용해 주세요.") 
        msg.add_alternative(html_body, subtype='html')

        with open(file_name, 'rb') as f:
            excel_data = f.read()
            
        msg.add_attachment(
            excel_data, 
            maintype='application', 
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            filename=file_name
        )

        if sender_email and app_password:
            try:
                with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                    smtp.login(sender_email, app_password)
                    smtp.send_message(msg)
                print("✅ 이메일 발송 완료!")
            except Exception as e:
                print(f"❌ 이메일 발송 실패: {e}")
        else:
            print("⚠️ GMAIL_USER 또는 GMAIL_APP_PASSWORD 환경변수가 설정되지 않아 메일을 발송하지 않았습니다.")

    except Exception as e: 
        print(f"\n❌ 최종 에러 발생: {e}")

if __name__ == "__main__":
    asyncio.run(main())
